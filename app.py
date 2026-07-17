import os
import secrets
import shutil
import io
import json
from datetime import datetime, timedelta
from functools import wraps
from flask import Flask, request, jsonify, g, send_file, send_from_directory
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, or_
import bcrypt
import jwt
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.utils import secure_filename
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger

load_dotenv()

app = Flask(__name__, static_folder='frontend', static_url_path='')
CORS(app)

# ===== CONFIGURAÇÕES =====
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'segredo_super_seguro_aconselhamento')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

# Configuração do banco de dados PostgreSQL (via DATABASE_URL)
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
if app.config['SQLALCHEMY_DATABASE_URI'] and app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgres://'):
    # Render fornece 'postgres://', mas SQLAlchemy exige 'postgresql://'
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Para compatibilidade com SQLite local (fallback)
if not app.config['SQLALCHEMY_DATABASE_URI']:
    # Se não houver DATABASE_URL, usar SQLite local (para desenvolvimento)
    DATABASE_PATH = './database/aconselhamento.db'
    os.makedirs(os.path.dirname(DATABASE_PATH), exist_ok=True)
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DATABASE_PATH}'

db = SQLAlchemy(app)

# Pastas para uploads e backups (persistência local – no Render free, serão efêmeras)
UPLOAD_FOLDER = './uploads'
BACKUP_FOLDER = './backups'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['BACKUP_FOLDER'] = BACKUP_FOLDER

PORT = int(os.getenv('PORT', 5000))
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'txt'}

# ===== MODELOS (TABELAS) =====
class Conselheiro(db.Model):
    __tablename__ = 'conselheiros'
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    senha_hash = db.Column(db.String(128), nullable=False)
    obs = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    pessoas = db.relationship('Pessoa', backref='conselheiro', lazy=True, cascade='all, delete-orphan')
    casais = db.relationship('Casal', backref='conselheiro', lazy=True, cascade='all, delete-orphan')
    sessoes = db.relationship('Sessao', backref='conselheiro', lazy=True, cascade='all, delete-orphan')
    anexos = db.relationship('Anexo', backref='conselheiro', lazy=True, cascade='all, delete-orphan')
    lembretes = db.relationship('Lembrete', backref='conselheiro', lazy=True, cascade='all, delete-orphan')
    logs = db.relationship('Log', backref='conselheiro', lazy=True, cascade='all, delete-orphan')

class Pessoa(db.Model):
    __tablename__ = 'pessoas'
    id = db.Column(db.Integer, primary_key=True)
    conselheiro_id = db.Column(db.Integer, db.ForeignKey('conselheiros.id', ondelete='CASCADE'), nullable=False)
    nome = db.Column(db.String(100), nullable=False)
    obs = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relacionamentos
    casais_homem = db.relationship('Casal', foreign_keys='Casal.id_homem', backref='homem', lazy=True, cascade='all, delete-orphan')
    casais_mulher = db.relationship('Casal', foreign_keys='Casal.id_mulher', backref='mulher', lazy=True, cascade='all, delete-orphan')

class Casal(db.Model):
    __tablename__ = 'casais'
    id = db.Column(db.Integer, primary_key=True)
    conselheiro_id = db.Column(db.Integer, db.ForeignKey('conselheiros.id', ondelete='CASCADE'), nullable=False)
    nome_casal = db.Column(db.String(100))
    id_homem = db.Column(db.Integer, db.ForeignKey('pessoas.id', ondelete='CASCADE'), nullable=False)
    id_mulher = db.Column(db.Integer, db.ForeignKey('pessoas.id', ondelete='CASCADE'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Sessao(db.Model):
    __tablename__ = 'sessoes'
    id = db.Column(db.Integer, primary_key=True)
    conselheiro_id = db.Column(db.Integer, db.ForeignKey('conselheiros.id', ondelete='CASCADE'), nullable=False)
    data = db.Column(db.String(20), nullable=False)  # formato YYYY-MM-DD
    caso_num = db.Column(db.String(20), nullable=False)
    sessao_num = db.Column(db.Integer, nullable=False)
    duracao = db.Column(db.String(20))
    is_casal = db.Column(db.Integer, nullable=False, default=0)  # 0 ou 1
    id_pessoa = db.Column(db.Integer, db.ForeignKey('pessoas.id', ondelete='SET NULL'))
    id_casal = db.Column(db.Integer, db.ForeignKey('casais.id', ondelete='SET NULL'))
    versiculo = db.Column(db.String(200))
    anotacao_sessao = db.Column(db.Text)
    tasks = db.Column(db.Text)  # JSON
    tarefas_anteriores = db.Column(db.Text)  # JSON
    assuntos_nesta = db.Column(db.Text)  # JSON
    assuntos_proximas = db.Column(db.Text)  # JSON
    status = db.Column(db.String(20), default='realizada')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Anexo(db.Model):
    __tablename__ = 'anexos'
    id = db.Column(db.Integer, primary_key=True)
    sessao_id = db.Column(db.Integer, db.ForeignKey('sessoes.id', ondelete='CASCADE'), nullable=False)
    conselheiro_id = db.Column(db.Integer, db.ForeignKey('conselheiros.id', ondelete='CASCADE'), nullable=False)
    nome_original = db.Column(db.String(200), nullable=False)
    nome_arquivo = db.Column(db.String(200), nullable=False)
    caminho = db.Column(db.String(300), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Lembrete(db.Model):
    __tablename__ = 'lembretes'
    id = db.Column(db.Integer, primary_key=True)
    conselheiro_id = db.Column(db.Integer, db.ForeignKey('conselheiros.id', ondelete='CASCADE'), nullable=False)
    sessao_id = db.Column(db.Integer, db.ForeignKey('sessoes.id', ondelete='CASCADE'))
    titulo = db.Column(db.String(100), nullable=False)
    descricao = db.Column(db.Text)
    data_lembrete = db.Column(db.String(20), nullable=False)  # YYYY-MM-DD
    concluido = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Log(db.Model):
    __tablename__ = 'logs'
    id = db.Column(db.Integer, primary_key=True)
    conselheiro_id = db.Column(db.Integer, db.ForeignKey('conselheiros.id', ondelete='CASCADE'), nullable=False)
    acao = db.Column(db.String(100), nullable=False)
    detalhes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class TokenBlacklist(db.Model):
    __tablename__ = 'tokens_blacklist'
    id = db.Column(db.Integer, primary_key=True)
    token = db.Column(db.String(500), nullable=False)
    expira_em = db.Column(db.DateTime, nullable=False)

class Configuracao(db.Model):
    __tablename__ = 'configuracoes'
    chave = db.Column(db.String(50), primary_key=True)
    valor = db.Column(db.String(200), nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# ===== CRIAÇÃO DAS TABELAS E MIGRAÇÕES =====
with app.app_context():
    db.create_all()
    # Verificar e adicionar colunas extras que podem não existir (para compatibilidade com migrações)
    # Como estamos usando SQLAlchemy, as colunas já estão definidas nos modelos.
    # Adicionar configurações padrão se não existirem
    configs_padrao = {
        'cor_primaria': '#b58b4b',
        'cor_secundaria': '#2d6a4f',
        'cor_fundo': '#f5f1eb',
        'cor_texto': '#2e2b28',
        'logo_url': '',
        'nome_igreja': 'Igreja Batista'
    }
    for chave, valor in configs_padrao.items():
        if not Configuracao.query.filter_by(chave=chave).first():
            db.session.add(Configuracao(chave=chave, valor=valor))
    db.session.commit()

# ===== SERVE FRONTEND =====
@app.route('/')
def serve_index():
    return send_file('frontend/index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('frontend', path)

# ===== BACKUP AUTOMÁTICO (não usa banco, apenas arquivo) =====
def fazer_backup():
    # Como agora usamos PostgreSQL, o backup do arquivo .db não é mais relevante.
    # Podemos manter para compatibilidade ou remover. Vou manter apenas como exemplo.
    # Na prática, o banco é gerenciado pelo Render.
    pass

# ===== AUTENTICAÇÃO =====
def gerar_token(conselheiro_id, refresh=False):
    exp = datetime.utcnow() + (timedelta(days=7) if refresh else timedelta(hours=1))
    payload = {'id': conselheiro_id, 'refresh': refresh, 'exp': exp}
    return jwt.encode(payload, app.config['SECRET_KEY'], algorithm='HS256')

def autenticar_token(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth_header = request.headers.get('Authorization')
        if not auth_header:
            return jsonify({'erro': 'Token não fornecido'}), 401
        try:
            token = auth_header.split(' ')[1]
            # Verificar blacklist
            if TokenBlacklist.query.filter_by(token=token).first():
                return jsonify({'erro': 'Token revogado'}), 403
            payload = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            if payload.get('refresh'):
                return jsonify({'erro': 'Use token de acesso'}), 403
            request.user_id = payload['id']
        except jwt.ExpiredSignatureError:
            return jsonify({'erro': 'Token expirado'}), 403
        except jwt.InvalidTokenError:
            return jsonify({'erro': 'Token inválido'}), 403
        return f(*args, **kwargs)
    return decorated

def log_acao(conselheiro_id, acao, detalhes=None):
    log = Log(conselheiro_id=conselheiro_id, acao=acao, detalhes=detalhes)
    db.session.add(log)
    db.session.commit()

def obter_proximo_sessao_num(conselheiro_id, caso_num):
    max_num = db.session.query(func.max(Sessao.sessao_num)).filter(
        Sessao.conselheiro_id == conselheiro_id,
        Sessao.caso_num == caso_num
    ).scalar()
    return (max_num or 0) + 1

# ===== ROTAS PÚBLICAS =====
@app.route('/api/registrar', methods=['POST'])
def registrar():
    data = request.json
    id_ = data.get('id')
    nome = data.get('nome')
    senha = data.get('senha')
    obs = data.get('obs', '')
    if not id_ or not nome or not senha:
        return jsonify({'erro': 'ID, nome e senha são obrigatórios'}), 400
    if not isinstance(id_, int) or id_ <= 0:
        return jsonify({'erro': 'ID deve ser um número inteiro positivo'}), 400
    if Conselheiro.query.get(id_):
        return jsonify({'erro': 'Este ID já está em uso'}), 400
    senha_hash = bcrypt.hashpw(senha.encode('utf-8'), bcrypt.gensalt(12)).decode('utf-8')
    cons = Conselheiro(id=id_, nome=nome, senha_hash=senha_hash, obs=obs)
    db.session.add(cons)
    db.session.commit()
    log_acao(id_, 'registro', f'Conselheiro {nome} cadastrado')
    return jsonify({'mensagem': 'Conselheiro registrado com sucesso'}), 201

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    id_ = data.get('id')
    senha = data.get('senha')
    if not id_ or not senha:
        return jsonify({'erro': 'ID e senha são obrigatórios'}), 400
    conselheiro = Conselheiro.query.get(id_)
    if not conselheiro:
        return jsonify({'erro': 'ID ou senha inválidos'}), 401
    if not bcrypt.checkpw(senha.encode('utf-8'), conselheiro.senha_hash.encode('utf-8')):
        return jsonify({'erro': 'ID ou senha inválidos'}), 401
    token = gerar_token(conselheiro.id, refresh=False)
    refresh = gerar_token(conselheiro.id, refresh=True)
    log_acao(conselheiro.id, 'login', 'Login realizado')
    return jsonify({
        'token': token,
        'refresh': refresh,
        'conselheiro': {'id': conselheiro.id, 'nome': conselheiro.nome, 'obs': conselheiro.obs}
    })

@app.route('/api/refresh', methods=['POST'])
def refresh_token():
    data = request.json
    refresh_token = data.get('refresh')
    if not refresh_token:
        return jsonify({'erro': 'Refresh token não fornecido'}), 401
    try:
        payload = jwt.decode(refresh_token, app.config['SECRET_KEY'], algorithms=['HS256'])
        if not payload.get('refresh'):
            return jsonify({'erro': 'Token inválido'}), 403
        if TokenBlacklist.query.filter_by(token=refresh_token).first():
            return jsonify({'erro': 'Token revogado'}), 403
        novo_token = gerar_token(payload['id'], refresh=False)
        return jsonify({'token': novo_token})
    except jwt.ExpiredSignatureError:
        return jsonify({'erro': 'Refresh token expirado'}), 403
    except jwt.InvalidTokenError:
        return jsonify({'erro': 'Refresh token inválido'}), 403

@app.route('/api/logout', methods=['POST'])
@autenticar_token
def logout():
    token = request.headers.get('Authorization').split(' ')[1]
    exp = datetime.utcnow() + timedelta(days=1)
    blacklist = TokenBlacklist(token=token, expira_em=exp)
    db.session.add(blacklist)
    db.session.commit()
    log_acao(request.user_id, 'logout', 'Logout realizado')
    return jsonify({'mensagem': 'Deslogado com sucesso'})

# ===== CONFIGURAÇÕES =====
@app.route('/api/configuracoes', methods=['GET'])
@autenticar_token
def get_configuracoes():
    configs = Configuracao.query.all()
    return jsonify({c.chave: c.valor for c in configs})

@app.route('/api/configuracoes', methods=['POST'])
@autenticar_token
def update_configuracoes():
    data = request.json
    for chave, valor in data.items():
        config = Configuracao.query.get(chave)
        if config:
            config.valor = valor
        else:
            db.session.add(Configuracao(chave=chave, valor=valor))
    db.session.commit()
    log_acao(request.user_id, 'atualizar_configuracoes', 'Configurações atualizadas')
    return jsonify({'mensagem': 'Configurações salvas'})

# ===== PESSOAS =====
@app.route('/api/pessoas', methods=['GET'])
@autenticar_token
def listar_pessoas():
    pessoas = Pessoa.query.filter_by(conselheiro_id=request.user_id).order_by(Pessoa.nome).all()
    return jsonify([{'id': p.id, 'nome': p.nome, 'obs': p.obs, 'conselheiro_id': p.conselheiro_id,
                     'created_at': p.created_at.isoformat() if p.created_at else None,
                     'updated_at': p.updated_at.isoformat() if p.updated_at else None} for p in pessoas])

@app.route('/api/pessoas', methods=['POST'])
@autenticar_token
def criar_pessoa():
    data = request.json
    nome = data.get('nome')
    obs = data.get('obs', '')
    if not nome:
        return jsonify({'erro': 'Nome é obrigatório'}), 400
    pessoa = Pessoa(conselheiro_id=request.user_id, nome=nome, obs=obs)
    db.session.add(pessoa)
    db.session.commit()
    log_acao(request.user_id, 'criar_pessoa', f'Pessoa {nome} criada')
    return jsonify({'id': pessoa.id, 'nome': pessoa.nome, 'obs': pessoa.obs}), 201

@app.route('/api/pessoas/<int:id>', methods=['DELETE'])
@autenticar_token
def deletar_pessoa(id):
    pessoa = Pessoa.query.filter_by(id=id, conselheiro_id=request.user_id).first()
    if not pessoa:
        return jsonify({'erro': 'Pessoa não encontrada'}), 404
    db.session.delete(pessoa)
    db.session.commit()
    return jsonify({'deletado': 1})

# ===== CASAIS =====
@app.route('/api/casais', methods=['GET'])
@autenticar_token
def listar_casais():
    casais = Casal.query.filter_by(conselheiro_id=request.user_id).order_by(Casal.nome_casal).all()
    return jsonify([{'id': c.id, 'nome_casal': c.nome_casal, 'id_homem': c.id_homem, 'id_mulher': c.id_mulher,
                     'conselheiro_id': c.conselheiro_id,
                     'created_at': c.created_at.isoformat() if c.created_at else None} for c in casais])

@app.route('/api/casais', methods=['POST'])
@autenticar_token
def criar_casal():
    data = request.json
    nome_casal = data.get('nome_casal', '')
    id_homem = data.get('id_homem')
    id_mulher = data.get('id_mulher')
    if not id_homem or not id_mulher:
        return jsonify({'erro': 'Selecione marido e esposa'}), 400
    # Verificar se as pessoas existem e pertencem ao conselheiro
    homem = Pessoa.query.filter_by(id=id_homem, conselheiro_id=request.user_id).first()
    mulher = Pessoa.query.filter_by(id=id_mulher, conselheiro_id=request.user_id).first()
    if not homem or not mulher:
        return jsonify({'erro': 'Pessoa não encontrada'}), 404
    casal = Casal(conselheiro_id=request.user_id, nome_casal=nome_casal, id_homem=id_homem, id_mulher=id_mulher)
    db.session.add(casal)
    db.session.commit()
    log_acao(request.user_id, 'criar_casal', f'Casal {nome_casal} criado')
    return jsonify({'id': casal.id}), 201

@app.route('/api/casais/<int:id>', methods=['DELETE'])
@autenticar_token
def deletar_casal(id):
    casal = Casal.query.filter_by(id=id, conselheiro_id=request.user_id).first()
    if not casal:
        return jsonify({'erro': 'Casal não encontrado'}), 404
    db.session.delete(casal)
    db.session.commit()
    return jsonify({'deletado': 1})

# ===== SESSÕES =====
@app.route('/api/sessoes', methods=['GET'])
@autenticar_token
def listar_sessoes():
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    offset = (page - 1) * per_page
    status_filter = request.args.get('status')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    assunto = request.args.get('assunto')

    query = Sessao.query.filter_by(conselheiro_id=request.user_id)
    if status_filter:
        query = query.filter(Sessao.status == status_filter)
    if data_inicio:
        query = query.filter(Sessao.data >= data_inicio)
    if data_fim:
        query = query.filter(Sessao.data <= data_fim)
    if assunto:
        query = query.filter(Sessao.assuntos_nesta.ilike(f'%{assunto}%'))

    total = query.count()
    rows = query.order_by(Sessao.data.desc()).offset(offset).limit(per_page).all()
    resultado = []
    for s in rows:
        d = {
            'id': s.id,
            'conselheiro_id': s.conselheiro_id,
            'data': s.data,
            'caso_num': s.caso_num,
            'sessao_num': s.sessao_num,
            'duracao': s.duracao,
            'is_casal': bool(s.is_casal),
            'id_pessoa': s.id_pessoa,
            'id_casal': s.id_casal,
            'versiculo': s.versiculo,
            'anotacao_sessao': s.anotacao_sessao,
            'tasks': json.loads(s.tasks) if s.tasks else [],
            'tarefas_anteriores': json.loads(s.tarefas_anteriores) if s.tarefas_anteriores else [],
            'assuntosNestaSessao': json.loads(s.assuntos_nesta) if s.assuntos_nesta else [],
            'assuntosProximasSessoes': json.loads(s.assuntos_proximas) if s.assuntos_proximas else [],
            'status': s.status,
            'created_at': s.created_at.isoformat() if s.created_at else None,
            'updated_at': s.updated_at.isoformat() if s.updated_at else None
        }
        resultado.append(d)

    return jsonify({
        'items': resultado,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page
    })

@app.route('/api/sessoes', methods=['POST'])
@autenticar_token
def criar_sessao():
    data = request.json
    data_sessao = data.get('data')
    caso_num = data.get('casoNum')
    if not data_sessao or not caso_num:
        return jsonify({'erro': 'Data e caso são obrigatórios'}), 400

    is_casal = data.get('isCasal', False)
    id_pessoa = data.get('idPessoa')
    id_casal = data.get('idCasal')
    if is_casal and not id_casal:
        return jsonify({'erro': 'Casal não selecionado'}), 400
    if not is_casal and not id_pessoa:
        return jsonify({'erro': 'Pessoa não selecionada'}), 400

    sessao_num = obter_proximo_sessao_num(request.user_id, caso_num)
    duracao = data.get('duracao', '')
    versiculo = data.get('versiculo', '')
    anotacao = data.get('anotacaoSessao', '')
    status = data.get('status', 'realizada')
    tasks = json.dumps(data.get('tasks', []))
    assuntos_nesta = json.dumps(data.get('assuntosNestaSessao', []))
    assuntos_proximas = json.dumps(data.get('assuntosProximasSessoes', []))

    tarefas_anteriores = []
    if 'tarefas_anteriores' in data and data['tarefas_anteriores']:
        tarefas_anteriores = data['tarefas_anteriores']
    else:
        # Buscar última sessão do mesmo caso
        ultima = Sessao.query.filter_by(conselheiro_id=request.user_id, caso_num=caso_num).order_by(Sessao.sessao_num.desc()).first()
        if ultima:
            tasks_ant = json.loads(ultima.tasks) if ultima.tasks else []
            for idx, t in enumerate(tasks_ant):
                if not t.get('avaliacao', '').strip():
                    tarefas_anteriores.append({
                        'sessao_origem_id': ultima.id,
                        'indice': idx,
                        'descricao': t.get('descricao', '')
                    })

    tarefas_anteriores_json = json.dumps(tarefas_anteriores)
    sessao = Sessao(
        conselheiro_id=request.user_id,
        data=data_sessao,
        caso_num=caso_num,
        sessao_num=sessao_num,
        duracao=duracao,
        is_casal=1 if is_casal else 0,
        id_pessoa=id_pessoa,
        id_casal=id_casal,
        versiculo=versiculo,
        anotacao_sessao=anotacao,
        tasks=tasks,
        tarefas_anteriores=tarefas_anteriores_json,
        assuntos_nesta=assuntos_nesta,
        assuntos_proximas=assuntos_proximas,
        status=status
    )
    db.session.add(sessao)
    db.session.commit()
    log_acao(request.user_id, 'criar_sessao', f'Sessão {caso_num}-{sessao_num} criada')
    return jsonify({'id': sessao.id, 'sessao_num': sessao_num}), 201

@app.route('/api/sessoes/<int:id>', methods=['PUT'])
@autenticar_token
def atualizar_sessao(id):
    sessao = Sessao.query.filter_by(id=id, conselheiro_id=request.user_id).first()
    if not sessao:
        return jsonify({'erro': 'Sessão não encontrada'}), 404

    data = request.json
    data_sessao = data.get('data')
    caso_num = data.get('casoNum')
    sessao_num = data.get('sessaoNum')
    if not data_sessao or not caso_num or sessao_num is None:
        return jsonify({'erro': 'Data, caso e número da sessão são obrigatórios'}), 400

    is_casal = data.get('isCasal', False)
    id_pessoa = data.get('idPessoa')
    id_casal = data.get('idCasal')
    if is_casal and not id_casal:
        return jsonify({'erro': 'Casal não selecionado'}), 400
    if not is_casal and not id_pessoa:
        return jsonify({'erro': 'Pessoa não selecionada'}), 400

    sessao.data = data_sessao
    sessao.caso_num = caso_num
    sessao.sessao_num = sessao_num
    sessao.duracao = data.get('duracao', '')
    sessao.is_casal = 1 if is_casal else 0
    sessao.id_pessoa = id_pessoa
    sessao.id_casal = id_casal
    sessao.versiculo = data.get('versiculo', '')
    sessao.anotacao_sessao = data.get('anotacaoSessao', '')
    sessao.status = data.get('status', 'realizada')
    sessao.tasks = json.dumps(data.get('tasks', []))
    sessao.tarefas_anteriores = json.dumps(data.get('tarefas_anteriores', []))
    sessao.assuntos_nesta = json.dumps(data.get('assuntosNestaSessao', []))
    sessao.assuntos_proximas = json.dumps(data.get('assuntosProximasSessoes', []))
    sessao.updated_at = datetime.utcnow()
    db.session.commit()
    log_acao(request.user_id, 'atualizar_sessao', f'Sessão {id} atualizada')
    return jsonify({'atualizado': 1})

@app.route('/api/sessoes/<int:id>', methods=['DELETE'])
@autenticar_token
def deletar_sessao(id):
    sessao = Sessao.query.filter_by(id=id, conselheiro_id=request.user_id).first()
    if not sessao:
        return jsonify({'erro': 'Sessão não encontrada'}), 404
    db.session.delete(sessao)
    db.session.commit()
    return jsonify({'deletado': 1})

# ===== ÚLTIMA SESSÃO (tarefas pendentes) =====
@app.route('/api/ultima_sessao', methods=['GET'])
@autenticar_token
def ultima_sessao():
    pessoa_id = request.args.get('pessoa_id', type=int)
    casal_id = request.args.get('casal_id', type=int)
    if not pessoa_id and not casal_id:
        return jsonify({'erro': 'Informe pessoa_id ou casal_id'}), 400

    if pessoa_id:
        ultima = Sessao.query.filter_by(conselheiro_id=request.user_id, id_pessoa=pessoa_id).order_by(Sessao.sessao_num.desc()).first()
    else:
        ultima = Sessao.query.filter_by(conselheiro_id=request.user_id, id_casal=casal_id).order_by(Sessao.sessao_num.desc()).first()

    if not ultima:
        return jsonify({'tarefas': []})

    tasks = json.loads(ultima.tasks) if ultima.tasks else []
    pendentes = []
    for idx, t in enumerate(tasks):
        if not t.get('avaliacao', '').strip():
            pendentes.append({
                'sessao_origem_id': ultima.id,
                'indice': idx,
                'descricao': t.get('descricao', '')
            })
    return jsonify({'tarefas': pendentes})

# ===== AVALIAR TAREFA ANTERIOR =====
@app.route('/api/avaliar_tarefa_anterior', methods=['POST'])
@autenticar_token
def avaliar_tarefa_anterior():
    data = request.json
    sessao_origem_id = data.get('sessao_origem_id')
    indice = data.get('indice')
    avaliacao = data.get('avaliacao', '')

    if sessao_origem_id is None or indice is None:
        return jsonify({'erro': 'Parâmetros inválidos'}), 400

    sessao = Sessao.query.filter_by(id=sessao_origem_id, conselheiro_id=request.user_id).first()
    if not sessao:
        return jsonify({'erro': 'Sessão não encontrada'}), 404

    tasks = json.loads(sessao.tasks) if sessao.tasks else []
    if indice >= len(tasks):
        return jsonify({'erro': 'Índice inválido'}), 400

    tasks[indice]['avaliacao'] = avaliacao
    sessao.tasks = json.dumps(tasks)
    db.session.commit()
    log_acao(request.user_id, 'avaliar_tarefa', f'Tarefa {indice} da sessão {sessao_origem_id} avaliada')
    return jsonify({'mensagem': 'Avaliação salva'})

# ===== ANEXOS =====
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/api/anexos/<int:sessao_id>', methods=['POST'])
@autenticar_token
def upload_anexo(sessao_id):
    if 'file' not in request.files:
        return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'erro': 'Nome de arquivo vazio'}), 400
    if not allowed_file(file.filename):
        return jsonify({'erro': 'Tipo de arquivo não permitido'}), 400

    sessao = Sessao.query.filter_by(id=sessao_id, conselheiro_id=request.user_id).first()
    if not sessao:
        return jsonify({'erro': 'Sessão não encontrada'}), 404

    filename = secure_filename(file.filename)
    unique_name = f"{secrets.token_hex(8)}_{filename}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
    file.save(filepath)

    anexo = Anexo(
        sessao_id=sessao_id,
        conselheiro_id=request.user_id,
        nome_original=filename,
        nome_arquivo=unique_name,
        caminho=filepath
    )
    db.session.add(anexo)
    db.session.commit()
    log_acao(request.user_id, 'upload_anexo', f'Anexo {filename} para sessão {sessao_id}')
    return jsonify({'id': anexo.id, 'mensagem': 'Arquivo enviado'}), 201

@app.route('/api/anexos/<int:sessao_id>', methods=['GET'])
@autenticar_token
def listar_anexos(sessao_id):
    anexos = Anexo.query.filter_by(sessao_id=sessao_id, conselheiro_id=request.user_id).all()
    return jsonify([{'id': a.id, 'nome_original': a.nome_original, 'nome_arquivo': a.nome_arquivo,
                     'caminho': a.caminho, 'created_at': a.created_at.isoformat() if a.created_at else None} for a in anexos])

@app.route('/api/anexos/<int:anexo_id>', methods=['DELETE'])
@autenticar_token
def deletar_anexo(anexo_id):
    anexo = Anexo.query.filter_by(id=anexo_id, conselheiro_id=request.user_id).first()
    if not anexo:
        return jsonify({'erro': 'Anexo não encontrado'}), 404
    if os.path.exists(anexo.caminho):
        os.remove(anexo.caminho)
    db.session.delete(anexo)
    db.session.commit()
    log_acao(request.user_id, 'deletar_anexo', f'Anexo {anexo_id} removido')
    return jsonify({'deletado': 1})

@app.route('/api/anexos/download/<int:anexo_id>', methods=['GET'])
@autenticar_token
def download_anexo(anexo_id):
    anexo = Anexo.query.filter_by(id=anexo_id, conselheiro_id=request.user_id).first()
    if not anexo:
        return jsonify({'erro': 'Anexo não encontrado'}), 404
    return send_file(anexo.caminho, download_name=anexo.nome_original)

# ===== LEMBRETES =====
@app.route('/api/lembretes', methods=['GET'])
@autenticar_token
def listar_lembretes():
    lembretes = Lembrete.query.filter_by(conselheiro_id=request.user_id).order_by(Lembrete.data_lembrete).all()
    return jsonify([{'id': l.id, 'titulo': l.titulo, 'descricao': l.descricao,
                     'data_lembrete': l.data_lembrete, 'concluido': l.concluido,
                     'sessao_id': l.sessao_id,
                     'created_at': l.created_at.isoformat() if l.created_at else None} for l in lembretes])

@app.route('/api/lembretes', methods=['POST'])
@autenticar_token
def criar_lembrete():
    data = request.json
    titulo = data.get('titulo')
    descricao = data.get('descricao', '')
    data_lembrete = data.get('data_lembrete')
    sessao_id = data.get('sessao_id')
    if not titulo or not data_lembrete:
        return jsonify({'erro': 'Título e data são obrigatórios'}), 400
    if sessao_id:
        if not Sessao.query.filter_by(id=sessao_id, conselheiro_id=request.user_id).first():
            return jsonify({'erro': 'Sessão não encontrada'}), 404
    lembrete = Lembrete(
        conselheiro_id=request.user_id,
        sessao_id=sessao_id,
        titulo=titulo,
        descricao=descricao,
        data_lembrete=data_lembrete
    )
    db.session.add(lembrete)
    db.session.commit()
    log_acao(request.user_id, 'criar_lembrete', f'Lembrete {titulo}')
    return jsonify({'id': lembrete.id}), 201

@app.route('/api/lembretes/<int:id>', methods=['PUT'])
@autenticar_token
def atualizar_lembrete(id):
    lembrete = Lembrete.query.filter_by(id=id, conselheiro_id=request.user_id).first()
    if not lembrete:
        return jsonify({'erro': 'Lembrete não encontrado'}), 404
    data = request.json
    concluido = data.get('concluido', 0)
    lembrete.concluido = concluido
    lembrete.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'atualizado': 1})

@app.route('/api/lembretes/<int:id>', methods=['DELETE'])
@autenticar_token
def deletar_lembrete(id):
    lembrete = Lembrete.query.filter_by(id=id, conselheiro_id=request.user_id).first()
    if not lembrete:
        return jsonify({'erro': 'Lembrete não encontrado'}), 404
    db.session.delete(lembrete)
    db.session.commit()
    return jsonify({'deletado': 1})

# ===== ESTATÍSTICAS =====
@app.route('/api/estatisticas', methods=['GET'])
@autenticar_token
def estatisticas():
    total_sessoes = Sessao.query.filter_by(conselheiro_id=request.user_id).count()
    total_pessoas = Pessoa.query.filter_by(conselheiro_id=request.user_id).count()
    total_casais = Casal.query.filter_by(conselheiro_id=request.user_id).count()
    # Tarefas pendentes: contar tasks onde avaliação é vazia (usando JSON functions do PostgreSQL)
    # Como é complexo com SQLAlchemy, podemos fazer uma query raw ou simplificar contando todas as tasks.
    # Vamos calcular via Python (mais fácil)
    sessoes = Sessao.query.filter_by(conselheiro_id=request.user_id).all()
    tarefas_pendentes = 0
    for s in sessoes:
        if s.tasks:
            tasks = json.loads(s.tasks)
            for t in tasks:
                if not t.get('avaliacao', '').strip():
                    tarefas_pendentes += 1
    lembretes_pendentes = Lembrete.query.filter_by(conselheiro_id=request.user_id, concluido=0).count()

    # Sessões por mês (últimos 12 meses)
    meses = db.session.query(
        func.strftime('%Y-%m', Sessao.data).label('mes'),
        func.count().label('total')
    ).filter(
        Sessao.conselheiro_id == request.user_id,
        Sessao.data >= (datetime.utcnow() - timedelta(days=365)).strftime('%Y-%m-%d')
    ).group_by('mes').order_by('mes').all()
    sessoes_por_mes = [{'mes': m.mes, 'total': m.total} for m in meses]

    return jsonify({
        'total_sessoes': total_sessoes,
        'total_pessoas': total_pessoas,
        'total_casais': total_casais,
        'tarefas_pendentes': tarefas_pendentes,
        'lembretes_pendentes': lembretes_pendentes,
        'sessoes_por_mes': sessoes_por_mes
    })

# ===== RELATÓRIOS =====
@app.route('/api/relatorios/tarefas-pendentes', methods=['GET'])
@autenticar_token
def relatorio_tarefas_pendentes():
    sessoes = Sessao.query.filter_by(conselheiro_id=request.user_id).filter(Sessao.tasks.isnot(None)).filter(Sessao.tasks != '[]').order_by(Sessao.data.desc()).all()
    resultado = []
    for s in sessoes:
        tasks = json.loads(s.tasks) if s.tasks else []
        pendentes = [{'descricao': t.get('descricao', ''), 'avaliacao': t.get('avaliacao', '')} for t in tasks if not t.get('avaliacao', '').strip()]
        if pendentes:
            resultado.append({
                'sessao_id': s.id,
                'caso': s.caso_num,
                'sessao': s.sessao_num,
                'data': s.data,
                'tarefas_pendentes': pendentes
            })
    return jsonify(resultado)

@app.route('/api/relatorios/sessoes-por-periodo', methods=['GET'])
@autenticar_token
def relatorio_sessoes_periodo():
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    if not data_inicio or not data_fim:
        return jsonify({'erro': 'Informe data_inicio e data_fim'}), 400
    sessoes = Sessao.query.filter_by(conselheiro_id=request.user_id).filter(Sessao.data.between(data_inicio, data_fim)).order_by(Sessao.data.desc()).all()
    resultado = []
    for s in sessoes:
        d = {
            'id': s.id,
            'data': s.data,
            'caso_num': s.caso_num,
            'sessao_num': s.sessao_num,
            'duracao': s.duracao,
            'is_casal': bool(s.is_casal),
            'id_pessoa': s.id_pessoa,
            'id_casal': s.id_casal,
            'versiculo': s.versiculo,
            'anotacao_sessao': s.anotacao_sessao,
            'tasks': json.loads(s.tasks) if s.tasks else [],
            'tarefas_anteriores': json.loads(s.tarefas_anteriores) if s.tarefas_anteriores else [],
            'assuntosNestaSessao': json.loads(s.assuntos_nesta) if s.assuntos_nesta else [],
            'assuntosProximasSessoes': json.loads(s.assuntos_proximas) if s.assuntos_proximas else [],
            'status': s.status
        }
        resultado.append(d)
    return jsonify(resultado)

@app.route('/api/relatorios/pessoas-atendidas', methods=['GET'])
@autenticar_token
def relatorio_pessoas_atendidas():
    # Pessoas que têm sessões
    pessoas = Pessoa.query.filter_by(conselheiro_id=request.user_id).all()
    resultado = []
    for p in pessoas:
        total_sessoes = Sessao.query.filter_by(conselheiro_id=request.user_id, id_pessoa=p.id).count()
        resultado.append({
            'id': p.id,
            'nome': p.nome,
            'obs': p.obs,
            'total_sessoes': total_sessoes
        })
    # Ordenar por total_sessoes decrescente
    resultado.sort(key=lambda x: x['total_sessoes'], reverse=True)
    return jsonify(resultado)

# ===== EXPORTAÇÃO EXCEL =====
@app.route('/api/exportar/excel', methods=['GET'])
@autenticar_token
def exportar_excel():
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    query = Sessao.query.filter_by(conselheiro_id=request.user_id)
    if data_inicio:
        query = query.filter(Sessao.data >= data_inicio)
    if data_fim:
        query = query.filter(Sessao.data <= data_fim)
    rows = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sessões'
    headers = ['ID', 'Data', 'Caso', 'Sessão', 'Duração', 'Tipo', 'Pessoa/Casal', 'Versículo', 'Anotação', 'Tarefas', 'Assuntos Nesta', 'Assuntos Próx', 'Status']
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    for s in rows:
        tipo = 'Casal' if s.is_casal else 'Individual'
        pessoa_casal = ''
        if s.is_casal and s.id_casal:
            casal = Casal.query.get(s.id_casal)
            pessoa_casal = casal.nome_casal if casal else ''
        elif s.id_pessoa:
            pessoa = Pessoa.query.get(s.id_pessoa)
            pessoa_casal = pessoa.nome if pessoa else ''
        tasks = json.loads(s.tasks) if s.tasks else []
        tasks_str = '; '.join([f"{t.get('descricao','')} [{t.get('avaliacao','')}]" for t in tasks])
        assuntos_nesta = json.loads(s.assuntos_nesta) if s.assuntos_nesta else []
        assuntos_proximas = json.loads(s.assuntos_proximas) if s.assuntos_proximas else []
        ws.append([
            s.id, s.data, s.caso_num, s.sessao_num, s.duracao,
            tipo, pessoa_casal, s.versiculo or '', s.anotacao_sessao or '',
            tasks_str, ', '.join(assuntos_nesta), ', '.join(assuntos_proximas), s.status or 'realizada'
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='sessoes.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ===== BACKUP MANUAL =====
@app.route('/api/backup', methods=['GET'])
@autenticar_token
def baixar_backup():
    # Com PostgreSQL, não temos um arquivo .db para baixar, mas podemos exportar dados.
    # Vamos retornar um arquivo JSON com todas as tabelas (simplificado).
    # Ou simplesmente retornamos uma mensagem informando que o backup é gerenciado pelo Render.
    return jsonify({'mensagem': 'Backup via PostgreSQL é gerenciado pelo Render. Utilize a ferramenta de backup do Render.'}), 200

# ===== LOGS =====
@app.route('/api/logs', methods=['GET'])
@autenticar_token
def listar_logs():
    logs = Log.query.filter_by(conselheiro_id=request.user_id).order_by(Log.created_at.desc()).limit(100).all()
    return jsonify([{'id': l.id, 'acao': l.acao, 'detalhes': l.detalhes, 'created_at': l.created_at.isoformat() if l.created_at else None} for l in logs])

# ===== PING =====
@app.route('/api/ping', methods=['GET'])
def ping():
    return jsonify({'status': 'ok', 'timestamp': datetime.utcnow().isoformat()})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=PORT, debug=False)